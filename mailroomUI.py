from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
from tkcalendar import Calendar
from tkcalendar import DateEntry
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import range_boundaries, get_column_letter
import datetime
import sys
import threading

#Initialising the root window for the frames and widgets and styling it accordingly
root = Tk()
root.geometry("1000x950")
style = ttk.Style()
style.theme_use("clam")
root.title("Mail Register")
root.configure(bg="grey28", pady=15)

#Creating tab control for implimenting both tabs
tabControl = ttk.Notebook(root)

#Creating tab for Incoming Mail and Out
incoming_Tab = Frame(tabControl, bg="grey30")
outgoing_Tab = Frame(tabControl, bg="grey30")

#Initializing and packing tabs
tabControl.add(incoming_Tab, text='Incoming Mail')
tabControl.add(outgoing_Tab, text='Outgoing Mail')
tabControl.pack(expand = 1, fill="both")

#Initialising the frames/containers where all the widgets will be placed in the Incoming tab
title = Label(incoming_Tab, borderwidth=3, text="Mail Register", font=('Ariel', 24), relief=GROOVE, bg="grey30", fg="cyan", padx=20)
title.pack(pady=15)

top_middle_frame = Frame(incoming_Tab, borderwidth=4, pady=30, padx=15, relief=SUNKEN, bg="grey28")
top_middle_frame.pack(pady=10)

middle_frame_incoming = Frame(incoming_Tab, borderwidth=4,relief=RIDGE, height=100, width=500, bg="grey30", padx=15, pady=15)
middle_frame_incoming.pack()

main_button_frame = Frame(incoming_Tab, borderwidth=4,relief=FLAT, bg="grey30", padx=15, pady=15)
main_button_frame.pack()

bottom_tablerows_frame = Frame(incoming_Tab,borderwidth=4,relief=RIDGE, bg="black", padx=15, pady=20, width=800, height=200)
bottom_tablerows_frame.pack()
bottom_tablerows_frame.grid_propagate(False)

insert_button_frame = Frame(main_button_frame, borderwidth=4,relief=RIDGE, bg="grey30", padx=15, pady=15)
insert_button_frame.pack(pady=15, side=LEFT)

clear_button_frameincoming = Frame(main_button_frame, borderwidth=4,relief=RIDGE, bg="grey30", padx=15, pady=15)
clear_button_frameincoming.pack(side=RIGHT)

bottom_tablerows_button = Frame(incoming_Tab, borderwidth=4,relief=RIDGE, bg="grey30", padx=15, pady=15)
bottom_tablerows_button.pack()

#Initialising the frames/containers where all the widgets will be placed in the Outgoing tab
title1 = Label(outgoing_Tab, borderwidth=3, text="Mail Register", font=('Ariel', 24), relief=GROOVE, bg="grey30", fg="cyan", padx=20)
title1.pack(pady=15)

top_middle_frame1 = Frame(outgoing_Tab, borderwidth=4, pady=30, padx=15, relief=SUNKEN, bg="grey28")
top_middle_frame1.pack(pady=10)

middle_frame_outgoing = Frame(outgoing_Tab, borderwidth=4,relief=RIDGE, height=100, width=500, bg="grey30", padx=15, pady=15)
middle_frame_outgoing.pack()

main_button_frame1 = Frame(outgoing_Tab, borderwidth=4,relief=FLAT, bg="grey30", padx=15, pady=15)
main_button_frame1.pack()

bottom_tablerows_frame1 = Frame(outgoing_Tab,borderwidth=4,relief=RIDGE, bg="black", padx=15, pady=20, width=800, height=200)
bottom_tablerows_frame1.pack()

insert_button_frame1 = Frame(main_button_frame1, borderwidth=4,relief=RIDGE, bg="grey30", padx=15, pady=15)
insert_button_frame1.pack(pady=15, side=LEFT)

clear_button_frameoutgoing = Frame(main_button_frame1, borderwidth=4,relief=RIDGE, bg="grey30", padx=15, pady=15)
clear_button_frameoutgoing.pack(side=RIGHT)



#Obtaining the current date as the default input for the calendar entry widget
current_date = datetime.datetime.now()
current_day = current_date.day
current_month = current_date.month
current_year = current_date.year

#Global variables, empty list initialisation for sheets and tables. Variables to store the data from the entry field widgets
filepath = None
inputString = StringVar()
inputString1 = StringVar()
inputString2 = StringVar()
inputString3 = StringVar()
inputString4 = StringVar()
inputString5 = StringVar()
inputString6 = StringVar()
inputString7 = StringVar()
inputString8 = StringVar()
inputString9 = StringVar()
inputString10 = StringVar()
inputString11 = StringVar()
tableList = []
sheetList = []

def loadWorkbook():
    global wb
    wb = load_workbook(filepath)
    
#Function to make the getentryIncoming function to run in a separate thread, keeping the GUI responsive and running the function as a separate process
#to increase the speed of exectuion.
def insert_incoming_thread():
    threading.Thread(target=getentryIncoming).start()

#Function to make the getentryOutgoing function to run in a separate thread, keeping the GUI responsive and running the function as a separate process
#to increase the speed of exectuion.
def insert_outgoing_thread():
    threading.Thread(target=getentryOutgoing).start()

#Function to run the insertintoTable function to run on a separate thread, keeping the GUI responsive and running the function as a separate process
#to increase the speed of execution
def inserting_table_data_thread(alist, worksheet, table, wb):
    thread = threading.Thread(target=insertintoTable, args=(alist, worksheet, table, wb))
    thread.start()

#Function to run the clearinputIncoming function to run on a separate thread, keeping the GUI responsive and running the function as a separate process
#to increase the speed of execution
def clear_input_incoming_thread():
    thread = threading.Thread(target=clearinputIncoming)
    thread.start()

#Function to run the clearinputOutgoing function to run on a separate thread, keeping the GUI responsive and running the function as a separate process
#to increase the speed of execution
def clear_input_outgoing_thread():
    thread = threading.Thread(target=clearinputOutgoing)
    thread.start()
    
def get_rows_thread():
    def callback():
        try:
            list_of_sheets = getSheets()
            list_of_tables = getTables()
            work_sheet = list_of_sheets[0]
            tab = list_of_tables[0]

            # Fetch the data
            data = fetchfiveRowsIncoming(work_sheet, tab)

            # Update the GUI in the main thread
            display_data_in_frame(data)

        except Exception as e:
            print(f"An error occurred: {e}")
            messagebox.showerror("ERROR", f"An error occurred: {e}")

    # Run the callback in the main thread
    root.after(0, callback)
    
#Function to browse local files with specified initial directory
def browseFiles():
    filename = filedialog.askopenfilename(title="Select a File",
                                           filetypes = (("Excel Files",
                                                         "*.xlsx*"),
                                                        ("all files",
                                                         "*.*")))
    if not filename.endswith('.xlsx'):
        messagebox.showerror("ERROR","File type is not '.xlsx'")

    else:
        pathlabel_top_middle_frame.configure(text="File Chosen: " + filename)
        pathlabel_top_middle_frame1.configure(text="File Chosen: " + filename)
        global filepath
        filepath = filename
        loadWorkbook()

#Function to clear all the input fields in the incoming mail tab.
def clearinputIncoming():
    courierlabelEntry.delete(0, END)
    fromlabelEntry.delete(0, END) 
    tolabelEntry.delete(0, END)
    descriptionlabelEntry.delete(0, END)

#Function to clear all the input fields in the outgoing mail tab
def clearinputOutgoing():
    courierlabelEntry1.delete(0, END)
    fromlabelEntry1.delete(0, END)
    tolabelEntry1.delete(0, END)
    descriptionlabelEntry1.delete(0, END)
    jobnumberlabelEntry.delete(0, END)
    consignmentlabelEntry.delete(0, END)
    
#Function to get the sheetnames from the excel file chosen by the user.
def getSheets():
    return [ws.title for ws in wb.worksheets]


#Function to get the tablenames from the excel file chosen by the user.
def getTables():
    tables = []
    for ws in wb.worksheets:
        for tbl in ws._tables:
            tables.append(tbl)
    return tables

#Function to get the number of empty entries from the input fields appended to a list. If all input fields are empty (depending on the amount there are)
#The function will return true and if it isn't it will return false. 
def getemptylistItems(alist, maxAmount):
    count = 0
    for the_list in alist:
            for items in the_list:
                if items == '':
                    count += 1

    if count >= maxAmount:
        return True
    else:
        return False

#Function that takes the list containing the data to input into the table, the excel sheet, the table itself and the work book and inserts
#The data into the table.
def insertintoTable(alist, worksheet, table, wb):
    try:
        # Error handling if worksheet or table is invalid
        if worksheet not in wb.sheetnames:
            raise ValueError(f"Worksheet '{worksheet}' not found in the workbook.")
        if table not in wb[worksheet].tables:
            raise ValueError(f"Table '{table}' not found in the worksheet '{worksheet}'.")

        ws = wb[worksheet]
        tbl = ws.tables[table]

        curr_ref = tbl.ref
        coord = list(range_boundaries(curr_ref))
        for row in alist:
            coord[-1] += 1
            ws.append(row)

        # Update the table reference to include new rows
        tbl.ref = f"{get_column_letter(coord[0])}{coord[1]}:{get_column_letter(coord[2])}{coord[3]}"

        # Align cells if necessary
        for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(alist[0])):
            for cell in col:
                cell.alignment = Alignment(horizontal='right')

        # Consider whether to save the workbook here or outside this function
        wb.save(filepath)

    except Exception as e:
        print(f"An error occurred: {e}")
        # Handle or re-raise the exception as needed


#Function to append all input into and array and to insert into the incoming mail register
def getentryIncoming():
    try:
        max_empties = 4
        date_Input = inputString.get()
        courier_Input = inputString1.get()
        from_Input = inputString2.get()
        to_Input = inputString3.get()
        description_Input = inputString4.get()

        list_of_sheets = getSheets()
        list_of_tables = getTables()

        data = [[date_Input, courier_Input, from_Input, to_Input, description_Input]]

        if getemptylistItems(data, max_empties) == True:
            messagebox.showerror("ERROR", "Input fields are empty")
        else:
            # Ensure that 'wb' is the loaded workbook object
            global wb
            inserting_table_data_thread(data, list_of_sheets[0], list_of_tables[0], wb)
            table_data = fetchfiveRowsIncoming(list_of_sheets[0], list_of_tables[0])
            display_data_in_frame(table_data)

    except TypeError as err:
        print("ERROR: ", err)
        messagebox.showerror("ERROR", "No path chosen/found")
    except PermissionError as err:
        print("ERROR: ", err)
        messagebox.showerror("ERROR", "File is open by another user")

#Function to append all input into and array and to insert into the outgoing mail register
def getentryOutgoing():
    try:
        max_empties = 6
        date_Input1 = inputString.get()
        courier_Input1 = inputString6.get()
        from_Input1 = inputString7.get()
        to_Input1 = inputString8.get()
        description_Input1 = inputString9.get()
        job_Number = inputString10.get()
        consignment_Note = inputString11.get()

        list_of_sheets = getSheets()
        list_of_tables = getTables()

        # Use the global workbook variable 'wb'
        global wb
        work_sheet = wb[list_of_sheets[1]]
        tab = work_sheet.tables[list_of_tables[1]]

        data = [[date_Input1, courier_Input1, from_Input1, to_Input1, description_Input1, job_Number, consignment_Note]]

        if getemptylistItems(data, max_empties) == True:
             messagebox.showerror("ERROR", "Input fields are empty")
        else:
            inserting_table_data_thread(data, list_of_sheets[1], list_of_tables[1], wb)

    except TypeError as err:
        print("No path entered/found: ", err)
        messagebox.showerror("ERROR", "No path chosen/found")
    except PermissionError as err:
        print("ERROR: ", err)
        messagebox.showerror("ERROR", "File is open by another user")


def fetchfiveRowsIncoming(sheet_name, table_name):
    # Load the workbook
    global wb
    # Get the specified sheet
    sheet = wb[sheet_name]
    
    # Get total number of rows in the sheet
    total_rows = sheet.max_row
    
    # Calculate the starting row index for fetching the last 5 rows
    start_row_index = max(total_rows - 4, 1)  # Ensure we don't go before the start of the sheet

    # Fetch the last 5 rows from the sheet
    data = []
    for row in sheet.iter_rows(min_row=start_row_index, max_row=total_rows, values_only=True):
        data.append(row)

    print(f"Start row index: {start_row_index}, End row: {total_rows}")
    print(f"Number of rows fetched: {len(data)}")
    return data

def display_data_in_frame(data):
    # Clear the frame before populating it with new data
    for widget in bottom_tablerows_frame.winfo_children():
        widget.destroy()
    
    # Loop through the fetched data and create labels to display each row
    for i, row in enumerate(data):
        for j, value in enumerate(row):
            label = Label(bottom_tablerows_frame, text=str(value), bg="black", fg="cyan", font=('Ariel', 14))
            label.grid(row=i, column=j, sticky="nsew")
            bottom_tablerows_frame.grid_columnconfigure(j, weight=1)

    bottom_tablerows_frame.grid_rowconfigure(len(data), weight=1)  

    

#Creating all entry field widgets with their lables for Incoming tab
dirpathButton = Button(top_middle_frame, text="Browse to Mail Register", borderwidth=2, bg="grey28", fg="cyan", command=browseFiles)
pathlabel_top_middle_frame = Label(top_middle_frame, width=50, font=('Ariel', 14), relief=SUNKEN, bg="white")


dateLabel = Label(middle_frame_incoming, relief=GROOVE, borderwidth=2, text="Date",bg="grey28", fg="cyan", width=10, height=2)
datelabelEntry = DateEntry(middle_frame_incoming, textvariable=inputString, borderwidth=2, bordercolor="black",
                           width=23,font=('Ariel', 21), year=current_year, month=current_month, day=current_day,
                           fg="white", relief=SUNKEN, bg="darkblue", date_pattern='dd/mm/y')

courierLabel = Label(middle_frame_incoming, text="Courier", relief=GROOVE,borderwidth=2, bg="grey28", fg="cyan", width=10, height=2)
courierlabelEntry = Entry(middle_frame_incoming, textvariable=inputString1, width=25, font=('Ariel', 21), relief=SUNKEN, bg="white")

fromLabel = Label(middle_frame_incoming, text="From", relief=GROOVE,borderwidth=2, bg="grey28", fg="cyan", width=10, height=2)
fromlabelEntry = Entry(middle_frame_incoming, textvariable=inputString2, width=25, font=('Ariel', 21), relief=SUNKEN, bg="white")

toLabel = Label(middle_frame_incoming, text="To", relief=GROOVE,borderwidth=2, bg="grey28", fg="cyan", width=10, height=2)
tolabelEntry = Entry(middle_frame_incoming, textvariable=inputString3, width=25, font=('Ariel', 21), relief=SUNKEN, bg="white")

descriptionLabel = Label(middle_frame_incoming, text="Desc.", relief=GROOVE,borderwidth=2, bg="grey28", fg="cyan", width=10, height=2)
descriptionlabelEntry = Entry(middle_frame_incoming, textvariable=inputString4, width=25, font=('Ariel', 21), relief=SUNKEN, bg="white")

clearButton = Button(clear_button_frameincoming, text="CLEAR", borderwidth=2, width=25, height=2, bg="grey28", fg="cyan", command=clear_input_incoming_thread)

insertButton = Button(insert_button_frame, text="INSERT", borderwidth=2, width=25, height=2, bg="grey28", fg="cyan", command=insert_incoming_thread)

print_rows_button = Button(bottom_tablerows_button, text="FETCH ROWS", borderwidth=2, width=25, height=2, bg="grey28", fg="cyan", command=get_rows_thread)



#Creating all entry field widgets with their lables for Outgoing tab
dirpathButton1 = Button(top_middle_frame1, text="Browse to Mail Register", borderwidth=2, bg="grey28", fg="cyan", command=browseFiles)
pathlabel_top_middle_frame1 = Label(top_middle_frame1, width=50, font=('Ariel', 14), relief=SUNKEN, bg="white")


dateLabel1 = Label(middle_frame_outgoing, relief=GROOVE, borderwidth=2, text="Date",bg="grey28", fg="cyan", width=10, height=2)
datelabelEntry1 = DateEntry(middle_frame_outgoing, textvariable=inputString5, borderwidth=2, bordercolor="black",
                           width=23,font=('Ariel', 21), year=current_year, month=current_month, day=current_day,
                           fg="white", relief=SUNKEN, bg="darkblue", date_pattern='dd/mm/y')

courierLabel1 = Label(middle_frame_outgoing, text="Courier", relief=GROOVE,borderwidth=2, bg="grey28", fg="cyan", width=10, height=2)
courierlabelEntry1 = Entry(middle_frame_outgoing, textvariable=inputString6, width=25, font=('Ariel', 21), relief=SUNKEN, bg="white")

fromLabel1 = Label(middle_frame_outgoing, text="From", relief=GROOVE,borderwidth=2, bg="grey28", fg="cyan", width=10, height=2)
fromlabelEntry1 = Entry(middle_frame_outgoing, textvariable=inputString7, width=25, font=('Ariel', 21), relief=SUNKEN, bg="white")

toLabel1 = Label(middle_frame_outgoing, text="To", relief=GROOVE,borderwidth=2, bg="grey28", fg="cyan", width=10, height=2)
tolabelEntry1 = Entry(middle_frame_outgoing, textvariable=inputString8, width=25, font=('Ariel', 21), relief=SUNKEN, bg="white")

descriptionLabel1 = Label(middle_frame_outgoing, text="Desc.", relief=GROOVE,borderwidth=2, bg="grey28", fg="cyan", width=10, height=2)
descriptionlabelEntry1 = Entry(middle_frame_outgoing, textvariable=inputString9, width=25, font=('Ariel', 21), relief=SUNKEN, bg="white")

jobnumberLabel = Label(middle_frame_outgoing, text="Job No.", relief=GROOVE,borderwidth=2, bg="grey28", fg="cyan", width=10, height=2)
jobnumberlabelEntry = Entry(middle_frame_outgoing, textvariable=inputString10, width=25, font=('Ariel', 21), relief=SUNKEN, bg="white")

consignmentLabel = Label(middle_frame_outgoing, text="Con Note", relief=GROOVE,borderwidth=2, bg="grey28", fg="cyan", width=10, height=2)
consignmentlabelEntry = Entry(middle_frame_outgoing, textvariable=inputString11, width=25, font=('Ariel', 21), relief=SUNKEN, bg="white")

clearButton1 = Button(clear_button_frameoutgoing, text="CLEAR", borderwidth=2, width=25, height=2, bg="grey28", fg="cyan", command=clear_input_outgoing_thread)

insertButton1 = Button(insert_button_frame1, text="INSERT", borderwidth=2, width=25, height=2, bg="grey28", fg="cyan", command=insert_outgoing_thread)

#Packing all widgets into a grid for Incoming tab
dirpathButton.grid(row=1, column=0)
pathlabel_top_middle_frame.grid(row=1, column=1)

dateLabel.grid(row=1, column=0, pady=25)
datelabelEntry.grid(row=1,column=1)

courierLabel.grid(row=1, column=2)
courierlabelEntry.grid(row=1, column=3)

fromLabel.grid(row=2, column=0)
fromlabelEntry.grid(row=2, column=1)

toLabel.grid(row=2, column=2)
tolabelEntry.grid(row=2, column=3)

descriptionLabel.grid(row=3, column=0, pady=25)
descriptionlabelEntry.grid(row=3, column=1)

clearButton.grid(row=3, column=2)

insertButton.grid(row=3, column=3)

print_rows_button.pack()

#Packing all widgets into a grid for Outgoing tab
dirpathButton1.grid(row=1, column=0)
pathlabel_top_middle_frame1.grid(row=1, column=1)

dateLabel1.grid(row=1, column=0, pady=25)
datelabelEntry1.grid(row=1,column=1)

courierLabel1.grid(row=1, column=2)
courierlabelEntry1.grid(row=1, column=3)

fromLabel1.grid(row=2, column=0)
fromlabelEntry1.grid(row=2, column=1)

toLabel1.grid(row=2, column=2)
tolabelEntry1.grid(row=2, column=3)

descriptionLabel1.grid(row=3, column=0, pady=25)
descriptionlabelEntry1.grid(row=3, column=1)

jobnumberLabel.grid(row=3, column=2)
jobnumberlabelEntry.grid(row=3, column=3)

consignmentLabel.grid(row=4, column=0)
consignmentlabelEntry.grid(row=4, column=1)

clearButton1.grid(row=4, column=2)

insertButton1.grid(row=4, column=3)

#Mainloop of tkinter window
root.mainloop()
